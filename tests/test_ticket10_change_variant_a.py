from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import pytest

pd = pytest.importorskip("pandas")
pytest.importorskip("openpyxl")

from openpyxl import Workbook, load_workbook

from tt_core.export.export_lp_copy import export_lp_copy_with_new_column
from tt_core.importers.import_service import ColumnMapping, import_asset
from tt_core.importers.xlsx_reader import read_tabular_data
from tt_core.jobs.job_service import run_change_variant_a_job
from tt_core.project.create_project import create_project, load_project_info
from tt_core.review.review_service import (
    list_changed_segments,
    list_proposals_for_asset,
    list_review_rows,
    list_segments,
    upsert_approved_translation,
)


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _import_change_asset(*, db_path: Path, project: object, xlsx_path: Path) -> str:
    file_bytes = xlsx_path.read_bytes()
    dataframe = read_tabular_data(
        file_type="xlsx",
        file_bytes=file_bytes,
        sheet_name="Sheet1",
    )
    summary = import_asset(
        db_path=db_path,
        project_id=project.project_id,
        source_locale=project.source_locale,
        dataframe=dataframe,
        file_type="xlsx",
        original_name=xlsx_path.name,
        column_mapping=ColumnMapping(
            mode="change_source_update",
            source_old="EN-OLD",
            source_new="EN-NEW",
            target="DE",
            target_locale="de-DE",
            key="Key",
        ),
        sheet_name="Sheet1",
        file_bytes=file_bytes,
        storage_path=str(xlsx_path),
        size_bytes=len(file_bytes),
    )
    return summary.asset_id


def test_change_variant_a_generates_changed_rows_only_and_exports_approved_rows(tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    created = create_project("Ticket 10 Change Fill", root=projects_root)
    project = load_project_info(created.slug, root=projects_root)

    source_xlsx = tmp_path / "change_variant_a.xlsx"
    _write_xlsx(
        source_xlsx,
        headers=["Key", "EN-OLD", "EN-NEW", "DE"],
        rows=[
            ["changed_one", "Attack", "Attack now", "Angriff"],
            ["unchanged", "Stay", "Stay", "Bleib"],
            ["changed_two", "Defend", "Defend now", "Verteidigen"],
        ],
    )

    asset_id = _import_change_asset(
        db_path=created.db_path,
        project=project,
        xlsx_path=source_xlsx,
    )

    changed_segments = list_changed_segments(db_path=created.db_path, asset_id=asset_id)
    assert [segment.key for segment in changed_segments] == ["changed_one", "changed_two"]

    result = run_change_variant_a_job(
        db_path=created.db_path,
        project_id=project.project_id,
        asset_id=asset_id,
        target_locale="de-DE",
        translator=lambda source_text, target_locale: f"[{target_locale}] {source_text}",
    )

    assert result.job_type == "change_variant_a"
    assert result.changed_segments == 2
    assert result.proposals_created == 2

    conn = sqlite3.connect(created.db_path)
    try:
        job_row = conn.execute(
            """
            SELECT job_type, decision_trace_json
            FROM jobs
            WHERE id = ?
            """,
            (result.job_id,),
        ).fetchone()
        assert job_row is not None
        assert job_row[0] == "change_variant_a"

        decision_trace = json.loads(str(job_row[1]))
        assert decision_trace["summary_counts"] == {
            "changed_rows": 2,
            "proposals_created": 2,
        }

        qa_rows = conn.execute(
            """
            SELECT s.key, q.type
            FROM qa_flags AS q
            INNER JOIN segments AS s
                ON s.id = q.segment_id
            WHERE s.asset_id = ?
              AND q.target_locale = 'de-DE'
            ORDER BY s.row_index, q.type
            """,
            (asset_id,),
        ).fetchall()
    finally:
        conn.close()

    assert qa_rows == [
        ("changed_one", "stale_source_change"),
        ("changed_two", "stale_source_change"),
    ]

    proposals = list_proposals_for_asset(
        db_path=created.db_path,
        asset_id=asset_id,
        target_locale="de-DE",
    )
    assert [proposal.candidate_type for proposal in proposals] == ["change_proposed", "change_proposed"]

    rows = list_review_rows(
        db_path=created.db_path,
        asset_id=asset_id,
        target_locale="de-DE",
    )
    rows_by_key = {row.key: row for row in rows}

    assert rows_by_key["changed_one"].baseline_text == "Angriff"
    assert rows_by_key["changed_one"].proposed_text == "[de-DE] Attack now"
    assert rows_by_key["changed_one"].proposed_type == "change_proposed"
    assert rows_by_key["unchanged"].baseline_text == "Bleib"
    assert rows_by_key["unchanged"].proposed_text is None

    approved_text = rows_by_key["changed_one"].proposed_text
    assert approved_text is not None
    upsert_approved_translation(
        db_path=created.db_path,
        segment_id=rows_by_key["changed_one"].segment_id,
        target_locale="de-DE",
        final_text=approved_text,
        approved_by="me",
    )

    export_result = export_lp_copy_with_new_column(
        db_path=created.db_path,
        project_slug=project.slug,
        project_path=project.project_path,
        asset_id=asset_id,
        target_locale="de-DE",
    )

    row_by_key = {segment.key: segment.row_index for segment in list_segments(db_path=created.db_path, asset_id=asset_id)}
    workbook = load_workbook(export_result.path)
    try:
        worksheet = workbook["Sheet1"]
        headers = [
            worksheet.cell(row=1, column=index).value
            for index in range(1, worksheet.max_column + 1)
        ]
        assert "NEW DE" in headers
        new_de_column = headers.index("NEW DE") + 1

        assert worksheet.cell(row=row_by_key["changed_one"], column=new_de_column).value == approved_text
        assert worksheet.cell(row=row_by_key["unchanged"], column=new_de_column).value is None
        assert worksheet.cell(row=row_by_key["changed_two"], column=new_de_column).value is None
    finally:
        workbook.close()
