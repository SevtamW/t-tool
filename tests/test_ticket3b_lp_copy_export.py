from __future__ import annotations

from pathlib import Path

import pytest

pd = pytest.importorskip("pandas")
pytest.importorskip("openpyxl")

from openpyxl import Workbook, load_workbook

from tt_core.export.export_lp_copy import export_lp_copy_with_new_column
from tt_core.importers.import_service import ColumnMapping, import_asset
from tt_core.importers.xlsx_reader import read_tabular_data
from tt_core.project.create_project import create_project, load_project_info
from tt_core.review.review_service import list_segments, upsert_approved_translation


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]], sheet_name: str = "Sheet1") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _import_local_xlsx(
    *,
    db_path: Path,
    project_id: str,
    source_locale: str,
    xlsx_path: Path,
    sheet_name: str = "Sheet1",
) -> str:
    file_bytes = xlsx_path.read_bytes()
    dataframe = read_tabular_data(
        file_type="xlsx",
        file_bytes=file_bytes,
        sheet_name=sheet_name,
    )

    summary = import_asset(
        db_path=db_path,
        project_id=project_id,
        source_locale=source_locale,
        dataframe=dataframe,
        file_type="xlsx",
        original_name=xlsx_path.name,
        column_mapping=ColumnMapping(
            source="EN",
            target="DE",
            cn=None,
            key=None,
            char_limit=None,
            context=[],
        ),
        sheet_name=sheet_name,
        file_bytes=file_bytes,
        storage_path=str(xlsx_path),
        size_bytes=len(file_bytes),
    )
    return summary.asset_id


def test_export_lp_copy_adds_new_column_and_writes_only_approved_rows(tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    created = create_project("LP Copy Export", root=projects_root)
    project = load_project_info(created.slug, root=projects_root)

    source_xlsx = tmp_path / "source.xlsx"
    _write_xlsx(
        source_xlsx,
        headers=["EN", "DE"],
        rows=[
            ["Hello", "Hallo"],
            ["Bye", "Tschuess"],
            ["Start", "Start"],
        ],
    )

    asset_id = _import_local_xlsx(
        db_path=created.db_path,
        project_id=project.project_id,
        source_locale=project.source_locale,
        xlsx_path=source_xlsx,
    )

    segments = list_segments(db_path=created.db_path, asset_id=asset_id)
    rows_by_source = {segment.source_text: segment.row_index for segment in segments}
    id_by_source = {segment.source_text: segment.id for segment in segments}

    upsert_approved_translation(
        db_path=created.db_path,
        segment_id=id_by_source["Hello"],
        target_locale="de-DE",
        final_text="Hallo Neu",
        approved_by="me",
    )
    upsert_approved_translation(
        db_path=created.db_path,
        segment_id=id_by_source["Start"],
        target_locale="de-DE",
        final_text="Start Neu",
        approved_by="me",
    )

    export_result = export_lp_copy_with_new_column(
        db_path=created.db_path,
        project_slug=project.slug,
        project_path=project.project_path,
        asset_id=asset_id,
        target_locale="de-DE",
    )

    assert export_result.path.exists()
    assert export_result.path.parent == project.project_path / "exports"
    assert export_result.path.name.startswith(f"lp_{project.slug}_{asset_id[:8]}_NEWDE_")
    assert export_result.row_count == 2

    original_wb = load_workbook(source_xlsx)
    try:
        original_ws = original_wb["Sheet1"]
        original_headers = [original_ws.cell(row=1, column=index).value for index in range(1, 3)]
        assert original_headers == ["EN", "DE"]
    finally:
        original_wb.close()

    exported_wb = load_workbook(export_result.path)
    try:
        ws = exported_wb["Sheet1"]
        headers = [ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)]
        assert headers.count("NEW DE") == 1
        new_de_column_index = headers.index("NEW DE") + 1

        assert ws.cell(row=rows_by_source["Hello"], column=new_de_column_index).value == "Hallo Neu"
        assert ws.cell(row=rows_by_source["Start"], column=new_de_column_index).value == "Start Neu"
        assert ws.cell(row=rows_by_source["Bye"], column=new_de_column_index).value is None
    finally:
        exported_wb.close()


def test_export_lp_copy_reuses_existing_new_column(tmp_path: Path) -> None:
    projects_root = tmp_path / "projects"
    created = create_project("LP Copy Existing Column", root=projects_root)
    project = load_project_info(created.slug, root=projects_root)

    source_xlsx = tmp_path / "source_existing_new_col.xlsx"
    _write_xlsx(
        source_xlsx,
        headers=["EN", "DE", "NEW DE"],
        rows=[
            ["Hello", "Hallo", "Old 1"],
            ["Bye", "Tschuess", "Old 2"],
            ["Start", "Start", "Old 3"],
        ],
    )

    asset_id = _import_local_xlsx(
        db_path=created.db_path,
        project_id=project.project_id,
        source_locale=project.source_locale,
        xlsx_path=source_xlsx,
    )

    segments = list_segments(db_path=created.db_path, asset_id=asset_id)
    rows_by_source = {segment.source_text: segment.row_index for segment in segments}
    id_by_source = {segment.source_text: segment.id for segment in segments}

    upsert_approved_translation(
        db_path=created.db_path,
        segment_id=id_by_source["Bye"],
        target_locale="de-DE",
        final_text="Neu 2",
        approved_by="me",
    )

    export_result = export_lp_copy_with_new_column(
        db_path=created.db_path,
        project_slug=project.slug,
        project_path=project.project_path,
        asset_id=asset_id,
        target_locale="de-DE",
    )

    exported_wb = load_workbook(export_result.path)
    try:
        ws = exported_wb["Sheet1"]
        headers = [ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)]
        assert headers.count("NEW DE") == 1
        assert ws.max_column == 3

        new_de_column_index = headers.index("NEW DE") + 1
        assert ws.cell(row=rows_by_source["Bye"], column=new_de_column_index).value == "Neu 2"
        assert ws.cell(row=rows_by_source["Hello"], column=new_de_column_index).value == "Old 1"
        assert ws.cell(row=rows_by_source["Start"], column=new_de_column_index).value == "Old 3"
    finally:
        exported_wb.close()
