from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import pytest

pd = pytest.importorskip("pandas")
pytest.importorskip("openpyxl")

from openpyxl import Workbook, load_workbook

from tt_core.export.export_lp_copy import export_lp_copy_with_new_column
from tt_core.importers.import_service import ColumnMapping, import_asset
from tt_core.importers.xlsx_reader import read_tabular_data
from tt_core.jobs.job_service import classify_change, run_change_variant_b_job
from tt_core.project.create_project import create_project, load_project_info
from tt_core.review.review_service import list_review_rows, list_segments, upsert_approved_translation


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _import_change_asset(*, db_path: Path, project: object, xlsx_path: Path) -> str:
    file_bytes = xlsx_path.read_bytes()
    dataframe = read_tabular_data(
        file_type="xlsx",
        file_bytes=file_bytes,
        sheet_name="Sheet1",
    )
    summary = import_asset(
        db_path=db_path,
        project_id=project.project_id,
        source_locale=project.source_locale,
        dataframe=dataframe,
        file_type="xlsx",
        original_name=xlsx_path.name,
        column_mapping=ColumnMapping(
            mode="change_source_update",
            source_old="EN-OLD",
            source_new="EN-NEW",
            target="DE",
            target_locale="de-DE",
            key="Key",
        ),
        sheet_name="Sheet1",
        file_bytes=file_bytes,
        storage_path=str(xlsx_path),
        size_bytes=len(file_bytes),
    )
    return summary.asset_id


def test_classify_change_uses_keep_update_and_flag_rules() -> None:
    keep = classify_change("Hello!", "Hello.")
    update = classify_change("Attack", "Attack right now")
    flagged = classify_change("Use {0}", "Use {1}")

    assert keep.decision == "KEEP"
    assert update.decision == "UPDATE"
    assert flagged.decision == "FLAG"


def test_change_variant_b_workflow_generates_review_data_and_exports_only_approved_rows(
    tmp_path: Path,
) -> None:
    projects_root = tmp_path / "projects"
    created = create_project("Ticket 9 Change Review", root=projects_root)
    project = load_project_info(created.slug, root=projects_root)

    source_xlsx = tmp_path / "change_variant_b.xlsx"
    _write_xlsx(
        source_xlsx,
        headers=["Key", "EN-OLD", "EN-NEW", "DE"],
        rows=[
            ["keep_punct", "Hello!", "Hello.", "Hallo!"],
            ["update_len", "Attack", "Attack right now", "Angriff"],
            ["flag_placeholder", "Use {0}", "Use {1}", "Nutze {0}"],
            ["unchanged", "Stay", "Stay", "Bleib"],
        ],
    )

    asset_id = _import_change_asset(
        db_path=created.db_path,
        project=project,
        xlsx_path=source_xlsx,
    )

    segments = list_segments(db_path=created.db_path, asset_id=asset_id)
    assert [segment.source_text_old for segment in segments] == [
        "Hello!",
        "Attack",
        "Use {0}",
        "Stay",
    ]
    assert [segment.source_text for segment in segments] == [
        "Hello.",
        "Attack right now",
        "Use {1}",
        "Stay",
    ]

    result = run_change_variant_b_job(
        db_path=created.db_path,
        project_id=project.project_id,
        asset_id=asset_id,
        target_locale="de-DE",
        translator=lambda source_text, target_locale: f"[{target_locale}] {source_text}",
    )

    assert result.job_type == "change_variant_b"
    assert result.changed_segments == 3
    assert result.keep_count == 1
    assert result.update_count == 1
    assert result.flag_count == 1

    conn = sqlite3.connect(created.db_path)
    try:
        job_row = conn.execute(
            """
            SELECT job_type, decision_trace_json
            FROM jobs
            WHERE id = ?
            """,
            (result.job_id,),
        ).fetchone()
        assert job_row is not None
        assert job_row[0] == "change_variant_b"

        decision_trace = json.loads(str(job_row[1]))
        assert decision_trace["summary_counts"] == {
            "changed_rows": 3,
            "keep": 1,
            "update": 1,
            "flag": 1,
        }

        qa_rows = conn.execute(
            """
            SELECT s.key, q.type
            FROM qa_flags AS q
            INNER JOIN segments AS s
                ON s.id = q.segment_id
            WHERE s.asset_id = ?
              AND q.target_locale = 'de-DE'
            ORDER BY s.row_index, q.type
            """,
            (asset_id,),
        ).fetchall()
        assert qa_rows == [
            ("keep_punct", "stale_source_change"),
            ("update_len", "stale_source_change"),
            ("flag_placeholder", "impact_flagged"),
            ("flag_placeholder", "stale_source_change"),
        ]

        candidate_rows = conn.execute(
            """
            SELECT s.key, tc.candidate_type
            FROM translation_candidates AS tc
            INNER JOIN segments AS s
                ON s.id = tc.segment_id
            WHERE s.asset_id = ?
              AND tc.target_locale = 'de-DE'
            ORDER BY s.row_index, tc.generated_at, tc.id
            """,
            (asset_id,),
        ).fetchall()
    finally:
        conn.close()

    assert candidate_rows.count(("update_len", "change_proposed")) == 1
    assert [row for row in candidate_rows if row[1] == "change_proposed"] == [
        ("update_len", "change_proposed")
    ]

    rows = list_review_rows(
        db_path=created.db_path,
        asset_id=asset_id,
        target_locale="de-DE",
    )
    rows_by_key = {row.key: row for row in rows}

    keep_row = rows_by_key["keep_punct"]
    assert keep_row.is_changed is True
    assert keep_row.baseline_text == "Hallo!"
    assert keep_row.proposed_text is None
    assert keep_row.change_decision == "KEEP"

    update_row = rows_by_key["update_len"]
    assert update_row.baseline_text == "Angriff"
    assert update_row.proposed_type == "change_proposed"
    assert update_row.proposed_text == "[de-DE] Attack right now"
    assert update_row.change_decision == "UPDATE"

    flag_row = rows_by_key["flag_placeholder"]
    assert flag_row.proposed_text is None
    assert flag_row.change_decision == "FLAG"

    approved_text = update_row.proposed_text
    assert approved_text is not None
    upsert_approved_translation(
        db_path=created.db_path,
        segment_id=update_row.segment_id,
        target_locale="de-DE",
        final_text=approved_text,
        approved_by="me",
    )

    conn = sqlite3.connect(created.db_path)
    try:
        tm_row = conn.execute(
            """
            SELECT source_text, target_text
            FROM tm_entries
            WHERE project_id = ?
              AND source_locale = ?
              AND target_locale = 'de-DE'
              AND source_text = 'Attack right now'
            ORDER BY created_at DESC, id DESC
            LIMIT 1
            """,
            (project.project_id, project.source_locale),
        ).fetchone()
        assert tm_row is not None
        assert tm_row[0] == "Attack right now"
        assert tm_row[1] == approved_text
    finally:
        conn.close()

    export_result = export_lp_copy_with_new_column(
        db_path=created.db_path,
        project_slug=project.slug,
        project_path=project.project_path,
        asset_id=asset_id,
        target_locale="de-DE",
    )

    row_by_key = {segment.key: segment.row_index for segment in segments}
    workbook = load_workbook(export_result.path)
    try:
        worksheet = workbook["Sheet1"]
        headers = [
            worksheet.cell(row=1, column=index).value
            for index in range(1, worksheet.max_column + 1)
        ]
        assert "NEW DE" in headers
        new_de_column = headers.index("NEW DE") + 1

        assert worksheet.cell(row=row_by_key["update_len"], column=new_de_column).value == approved_text
        assert worksheet.cell(row=row_by_key["keep_punct"], column=new_de_column).value is None
        assert worksheet.cell(row=row_by_key["flag_placeholder"], column=new_de_column).value is None
        assert worksheet.cell(row=row_by_key["unchanged"], column=new_de_column).value is None
    finally:
        workbook.close()
