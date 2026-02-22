from __future__ import annotations

import json
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text

from tt_core.db.schema import initialize_database
from tt_core.review.review_service import ApprovedPatchRow, list_approved_for_asset

_SAFE_CHARS = re.compile(r"[^a-zA-Z0-9_-]+")


@dataclass(slots=True)
class LpCopyExportResult:
    path: Path
    row_count: int
    new_column_name: str
    source_path: Path
    warnings: list[str] = field(default_factory=list)


@dataclass(slots=True)
class _AssetExportInfo:
    id: str
    project_id: str
    asset_type: str
    original_name: str | None
    storage_path: str | None


def _utc_filename_timestamp() -> str:
    now = datetime.now(timezone.utc).replace(microsecond=0)
    return now.isoformat().replace("+00:00", "Z").replace(":", "-")


def _safe_fragment(value: str) -> str:
    cleaned = _SAFE_CHARS.sub("_", value.strip())
    return cleaned.strip("_") or "export"


def _locale_short(target_locale: str) -> str:
    base = target_locale.split("-", 1)[0].strip()
    if base:
        return base.upper()
    return target_locale.strip().upper() or "XX"


def _new_column_name(target_locale: str) -> str:
    return f"NEW {_locale_short(target_locale)}"


def _new_column_filename_token(target_locale: str) -> str:
    return f"NEW{_locale_short(target_locale)}"


def _import_copy_path(project_path: Path, asset_id: str, original_name: str | None) -> Path:
    safe_name = Path(original_name or f"{asset_id}.xlsx").name
    return Path(project_path) / "imports" / f"{asset_id}_{safe_name}"


def _load_asset_info(*, db_path: Path, asset_id: str) -> _AssetExportInfo:
    engine = initialize_database(Path(db_path))
    try:
        with engine.connect() as connection:
            row = connection.execute(
                text(
                    """
                    SELECT id, project_id, asset_type, original_name, storage_path
                    FROM assets
                    WHERE id = :asset_id
                    LIMIT 1
                    """
                ),
                {"asset_id": asset_id},
            ).first()
    finally:
        engine.dispose()

    if row is None:
        raise ValueError(f"Asset not found: {asset_id}")

    return _AssetExportInfo(
        id=str(row[0]),
        project_id=str(row[1]),
        asset_type=str(row[2]),
        original_name=row[3],
        storage_path=row[4],
    )


def _resolve_source_xlsx_path(
    *,
    project_path: Path,
    asset: _AssetExportInfo,
    uploaded_xlsx_bytes: bytes | None,
    store_uploaded_copy: bool,
) -> tuple[Path, list[str]]:
    warnings: list[str] = []

    if asset.storage_path:
        storage_candidate = Path(asset.storage_path).expanduser()
        if storage_candidate.is_file():
            return storage_candidate, warnings
        warnings.append(f"Storage path not found: {storage_candidate}")

    imports_candidate = _import_copy_path(project_path, asset.id, asset.original_name)
    if imports_candidate.is_file():
        return imports_candidate, warnings

    if store_uploaded_copy and uploaded_xlsx_bytes is not None:
        imports_candidate.parent.mkdir(parents=True, exist_ok=True)
        imports_candidate.write_bytes(uploaded_xlsx_bytes)
        warnings.append(f"Stored uploaded XLSX copy: {imports_candidate}")
        return imports_candidate, warnings

    raise FileNotFoundError("Original XLSX not available; use Patch Export instead.")


def _resolve_sheet_name(
    *,
    db_path: Path,
    project_id: str,
    fallback_sheet_name: str | None,
) -> tuple[str | None, list[str]]:
    warnings: list[str] = []
    engine = initialize_database(Path(db_path))
    try:
        with engine.connect() as connection:
            rows = connection.execute(
                text(
                    """
                    SELECT mapping_json
                    FROM schema_profiles
                    WHERE project_id = :project_id
                    ORDER BY updated_at DESC
                    """
                ),
                {"project_id": project_id},
            ).all()
    finally:
        engine.dispose()

    preferred = (fallback_sheet_name or "").strip() or None
    fallback_mapping_sheet: str | None = None

    for row in rows:
        try:
            payload = json.loads(row[0] or "{}")
        except (TypeError, ValueError):
            continue
        if not isinstance(payload, dict):
            continue
        if str(payload.get("file_type", "")).lower() != "xlsx":
            continue

        mapping_sheet = str(payload.get("sheet_name") or "").strip() or None
        if mapping_sheet is None:
            continue

        if preferred and mapping_sheet == preferred:
            return mapping_sheet, warnings
        if fallback_mapping_sheet is None:
            fallback_mapping_sheet = mapping_sheet

    if fallback_mapping_sheet:
        if preferred and fallback_mapping_sheet != preferred:
            warnings.append(
                f"Using schema profile sheet '{fallback_mapping_sheet}' instead of '{preferred}'."
            )
        return fallback_mapping_sheet, warnings

    return preferred, warnings


def _ensure_new_column(worksheet: Worksheet, new_column_name: str) -> int:
    for column_index in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=1, column=column_index).value
        if header_value is None:
            continue
        if str(header_value).strip() == new_column_name:
            return column_index

    new_column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=new_column_index).value = new_column_name
    return new_column_index


def _write_rows_for_sheet(
    *,
    worksheet: Worksheet,
    rows: list[ApprovedPatchRow],
    new_column_name: str,
) -> int:
    new_column_index = _ensure_new_column(worksheet, new_column_name)
    written_rows = 0
    for row in rows:
        if row.row_index is None or row.row_index < 2:
            continue
        worksheet.cell(row=row.row_index, column=new_column_index).value = row.approved_target_text
        written_rows += 1
    return written_rows


def export_lp_copy_with_new_column(
    *,
    db_path: Path,
    project_slug: str,
    project_path: Path,
    asset_id: str,
    target_locale: str,
    uploaded_xlsx_bytes: bytes | None = None,
    store_uploaded_copy: bool = False,
) -> LpCopyExportResult:
    asset = _load_asset_info(db_path=db_path, asset_id=asset_id)
    if asset.asset_type.lower() != "xlsx":
        raise ValueError("LP copy export supports only XLSX assets.")

    source_path, source_warnings = _resolve_source_xlsx_path(
        project_path=Path(project_path),
        asset=asset,
        uploaded_xlsx_bytes=uploaded_xlsx_bytes,
        store_uploaded_copy=store_uploaded_copy,
    )

    approved_rows = list_approved_for_asset(
        db_path=db_path,
        asset_id=asset_id,
        target_locale=target_locale,
    )
    if not approved_rows:
        raise ValueError("No approved translations found for this asset and locale")

    fallback_sheet = next((row.sheet_name for row in approved_rows if row.sheet_name), None)
    mapped_sheet_name, sheet_warnings = _resolve_sheet_name(
        db_path=db_path,
        project_id=asset.project_id,
        fallback_sheet_name=fallback_sheet,
    )

    exports_dir = Path(project_path) / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)

    filename = (
        f"lp_{_safe_fragment(project_slug)}_{asset_id[:8]}_"
        f"{_safe_fragment(_new_column_filename_token(target_locale))}_"
        f"{_utc_filename_timestamp()}.xlsx"
    )
    output_path = exports_dir / filename
    shutil.copy2(source_path, output_path)

    workbook = load_workbook(output_path)
    warnings = [*source_warnings, *sheet_warnings]
    new_col_name = _new_column_name(target_locale)

    try:
        default_sheet_name = mapped_sheet_name
        if default_sheet_name and default_sheet_name not in workbook.sheetnames:
            warnings.append(
                f"Sheet '{default_sheet_name}' not found; using active sheet '{workbook.active.title}'."
            )
            default_sheet_name = workbook.active.title
        if default_sheet_name is None:
            warnings.append(f"Sheet mapping not found; using active sheet '{workbook.active.title}'.")
            default_sheet_name = workbook.active.title

        rows_by_sheet: dict[str, list[ApprovedPatchRow]] = {}
        for row in approved_rows:
            sheet_name = (row.sheet_name or default_sheet_name or "").strip() or workbook.active.title
            rows_by_sheet.setdefault(sheet_name, []).append(row)

        written_rows = 0
        for sheet_name, sheet_rows in rows_by_sheet.items():
            if sheet_name not in workbook.sheetnames:
                warnings.append(f"Sheet '{sheet_name}' not found; skipped {len(sheet_rows)} approved row(s).")
                continue
            worksheet = workbook[sheet_name]
            written_rows += _write_rows_for_sheet(
                worksheet=worksheet,
                rows=sheet_rows,
                new_column_name=new_col_name,
            )

        workbook.save(output_path)
    finally:
        workbook.close()

    return LpCopyExportResult(
        path=output_path,
        row_count=written_rows,
        new_column_name=new_col_name,
        source_path=source_path,
        warnings=warnings,
    )

